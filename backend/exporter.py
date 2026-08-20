import os
import logging
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from schemas import DinatInvoiceDocument

logger = logging.getLogger("backend.exporter")

SHEET_COLUMNS = {
    "DATOS-CLIENTES": [
        "ID_COMPRA",
        "FECHA_VENTA",
        "RAZON_SOCIAL",
        "RTN_CLIENTE",
        "TIENDA",
        "COD_TIENDA",
        "CIUDAD",
        "DIRECCION",
        "LATITUD",
        "LONGITUD",
        "CONTACTO_CLIENTE"
    ],
    "PRODUCTO": [
        "ID_COMPRA",
        "CODIGO_PRODUCTO",
        "DESCRIPCION_PRODUCTO",
        "SABOR",
        "PRESENTACION",
        "CANTIDAD",
        "UNIDAD_TOTALES",
        "PRECIO",
        "IMPORTE"
    ],
    "ENTREGA": [
        "ID_COMPRA",
        "NOMBRE-TRASLADOR",
        "IDENTIDAD",
        "RUTA",
        "CODIOGO_EMPLEADO",
        "CARGO",
        "UNIDAD",
        "ESTADO_ENTREGA",
        "HORA_LLEGADA",
        "HORA_FINALIZACION",
        "OBSERVACION"
    ]
}


def resolve_safe_excel_path(file_path: str) -> str:
    """
    Validates target path permissions. If target path is root C:\\ or a non-writable/restricted path,
    resolves a safe, writable local user path.
    """
    clean_path = os.path.abspath(file_path)
    parent_dir = os.path.dirname(clean_path)
    
    if parent_dir == os.path.abspath(os.sep) or not os.path.exists(parent_dir):
        try:
            os.makedirs(parent_dir, exist_ok=True)
        except (PermissionError, OSError):
            fallback_dir = os.path.abspath(os.path.join(os.getcwd(), "exports"))
            os.makedirs(fallback_dir, exist_ok=True)
            filename = os.path.basename(file_path)
            clean_path = os.path.join(fallback_dir, filename if filename else "MATRIZ-ORDEN-COMPRA.xlsx")
            logger.warning(f"Target path '{file_path}' permission restricted. Saving safely to fallback path '{clean_path}'")

    return clean_path


def get_last_populated_row(ws, col_count: int = 15) -> int:
    """
    Scans from current ws.max_row downwards to find the exact last row that has non-empty values,
    preventing openpyxl from overwriting existing rows or skipping due to empty styled cells.
    """
    for r in range(ws.max_row, 0, -1):
        if any(ws.cell(row=r, column=c).value is not None for c in range(1, col_count + 1)):
            return r
    return 0


def format_header_and_widths(ws, columns: List[str]):
    """Applies header formatting and adjusts column widths."""
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)


def find_or_create_sheet(wb: openpyxl.Workbook, sheet_key: str, order_prefix: Optional[str] = None) -> openpyxl.worksheet.worksheet.Worksheet:
    """
    Finds existing sheet matching sheet_key or order_prefix-sheet_key, or creates it.
    """
    target_name = f"{order_prefix}-{sheet_key}" if order_prefix else sheet_key
    
    # Check exact match
    if target_name in wb.sheetnames:
        return wb[target_name]
    
    # Check if sheet ending with -sheet_key or exact sheet_key exists
    for name in wb.sheetnames:
        if name == sheet_key or name.endswith(f"-{sheet_key}"):
            return wb[name]
            
    # If default Sheet exists and is empty, reuse it
    if "Sheet" in wb.sheetnames and wb["Sheet"].max_row <= 1 and wb["Sheet"].cell(1, 1).value is None:
        ws = wb["Sheet"]
        ws.title = target_name
        ws.append(SHEET_COLUMNS[sheet_key])
        format_header_and_widths(ws, SHEET_COLUMNS[sheet_key])
        return ws

    # Otherwise create a new sheet
    ws = wb.create_sheet(title=target_name)
    ws.append(SHEET_COLUMNS[sheet_key])
    format_header_and_widths(ws, SHEET_COLUMNS[sheet_key])
    return ws


def generate_or_append_excel(documents: List[DinatInvoiceDocument], file_path: str, mode: str = "append") -> Dict[str, Any]:
    """
    Creates a new Excel workbook or appends invoice data across the 3 exact sheets
    matching MATRIZ-ORDEN-COMPRA.xlsx:
    1. DATOS-CLIENTES (11 columns)
    2. PRODUCTO (9 columns)
    3. ENTREGA (11 columns)
    """
    target_file = resolve_safe_excel_path(file_path)

    try:
        file_exists = os.path.exists(target_file)
        
        if mode == "create" or not file_exists:
            wb = openpyxl.Workbook()
            # Remove default sheet once we create our sheets
            default_sheet = wb.active
            order_prefix = documents[0].document_metadata.order_number if documents else None
            
            # Create 3 sheets
            ws_client = wb.create_sheet(title=f"{order_prefix}-DATOS-CLIENTES" if order_prefix else "DATOS-CLIENTES")
            ws_client.append(SHEET_COLUMNS["DATOS-CLIENTES"])
            
            ws_prod = wb.create_sheet(title=f"{order_prefix}-PRODUCTO" if order_prefix else "PRODUCTO")
            ws_prod.append(SHEET_COLUMNS["PRODUCTO"])
            
            ws_deliv = wb.create_sheet(title=f"{order_prefix}-ENTREGA" if order_prefix else "ENTREGA")
            ws_deliv.append(SHEET_COLUMNS["ENTREGA"])
            
            if default_sheet in wb.worksheets:
                wb.remove(default_sheet)
        else:
            wb = openpyxl.load_workbook(target_file)

        data_font = Font(name="Calibri", size=10)
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )

        total_product_rows_added = 0

        for doc in documents:
            order_no = doc.document_metadata.order_number
            date_formatted = doc.document_metadata.issue_date.replace("-", "/")
            
            prefix = order_no if mode == "create" else None

            # --- 1. SHEET: DATOS-CLIENTES ---
            ws_client = find_or_create_sheet(wb, "DATOS-CLIENTES", prefix)
            last_r_client = get_last_populated_row(ws_client, len(SHEET_COLUMNS["DATOS-CLIENTES"]))
            next_r_client = last_r_client + 1 if last_r_client > 0 else 2

            client_row = [
                order_no,
                date_formatted,
                doc.client.company_name,
                doc.client.rtn,
                doc.client.store_name,
                doc.client.store_code,
                doc.client.city_department,
                doc.client.address,
                doc.client.coordinates.latitude,
                doc.client.coordinates.longitude,
                doc.client.store_contact
            ]

            for col_idx, val in enumerate(client_row, start=1):
                cell = ws_client.cell(row=next_r_client, column=col_idx, value=val)
                cell.font = data_font
                cell.border = thin_border
                if col_idx in [9, 10]:  # Coordinates
                    cell.alignment = right_align
                    cell.number_format = '0.000000'
                elif col_idx in [1, 2, 4, 6]:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align

            # --- 2. SHEET: PRODUCTO ---
            ws_prod = find_or_create_sheet(wb, "PRODUCTO", prefix)
            last_r_prod = get_last_populated_row(ws_prod, len(SHEET_COLUMNS["PRODUCTO"]))
            next_r_prod = last_r_prod + 1 if last_r_prod > 0 else 2

            items = doc.items if doc.items else []
            for item in items:
                prod_row = [
                    order_no,
                    item.code,
                    item.description,
                    item.flavor,
                    item.package_type,
                    item.boxes_quantity,
                    item.total_units,
                    item.unit_price,
                    item.total_amount
                ]
                for col_idx, val in enumerate(prod_row, start=1):
                    cell = ws_prod.cell(row=next_r_prod, column=col_idx, value=val)
                    cell.font = data_font
                    cell.border = thin_border
                    if col_idx in [6, 7]:  # Cantidad, Unidades
                        cell.alignment = right_align
                        cell.number_format = '#,##0'
                    elif col_idx in [8, 9]:  # Precio, Importe
                        cell.alignment = right_align
                        cell.number_format = '#,##0.00'
                    elif col_idx in [1, 2]:
                        cell.alignment = center_align
                    else:
                        cell.alignment = left_align

                next_r_prod += 1
                total_product_rows_added += 1

            # --- 3. SHEET: ENTREGA ---
            ws_deliv = find_or_create_sheet(wb, "ENTREGA", prefix)
            last_r_deliv = get_last_populated_row(ws_deliv, len(SHEET_COLUMNS["ENTREGA"]))
            next_r_deliv = last_r_deliv + 1 if last_r_deliv > 0 else 2

            deliv_row = [
                order_no,
                doc.transport_logistics.driver_name,
                doc.transport_logistics.national_id,
                doc.transport_logistics.assigned_route,
                doc.transport_logistics.employee_id,
                doc.transport_logistics.role,
                doc.transport_logistics.transport_unit,
                doc.delivery_status.status,
                doc.delivery_status.arrival_time,
                doc.delivery_status.completion_time,
                doc.delivery_status.observations
            ]

            for col_idx, val in enumerate(deliv_row, start=1):
                cell = ws_deliv.cell(row=next_r_deliv, column=col_idx, value=val)
                cell.font = data_font
                cell.border = thin_border
                if col_idx in [1, 3, 5, 9, 10]:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align

        # Format headers and auto-adjust widths for all sheets in workbook
        for name in wb.sheetnames:
            ws = wb[name]
            sheet_type = "DATOS-CLIENTES" if "DATOS-CLIENTES" in name else "PRODUCTO" if "PRODUCTO" in name else "ENTREGA" if "ENTREGA" in name else None
            if sheet_type:
                format_header_and_widths(ws, SHEET_COLUMNS[sheet_type])

        try:
            wb.save(target_file)
        except PermissionError:
            fallback_file = os.path.abspath(os.path.join(os.getcwd(), "exports", f"MATRIZ_{os.path.basename(target_file)}"))
            os.makedirs(os.path.dirname(fallback_file), exist_ok=True)
            wb.save(fallback_file)
            target_file = fallback_file
            logger.warning(f"File locked or write permission denied for '{file_path}'. Saved to '{fallback_file}'")

        return {
            "success": True,
            "message": f"Successfully exported {total_product_rows_added} product row(s) and 3 matrix sheets to Excel.",
            "target_path": os.path.abspath(target_file),
            "mode": mode,
            "rows_added": total_product_rows_added
        }

    except Exception as e:
        logger.error(f"Error generating Excel file: {e}", exc_info=True)
        raise PermissionError(f"Permission restricted for file path '{file_path}'. Details: {e}")
