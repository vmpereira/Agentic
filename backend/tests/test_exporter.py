import os
import openpyxl
import pytest
from exporter import (
    generate_or_append_excel,
    resolve_safe_excel_path,
    get_last_populated_row,
    SHEET_COLUMNS,
)


def test_create_new_excel_three_sheets(tmp_path, mock_dinat_invoice_doc):
    excel_file = tmp_path / "test_matriz_create.xlsx"
    file_path = str(excel_file)

    res = generate_or_append_excel([mock_dinat_invoice_doc], file_path, mode="create")
    assert res["success"] is True
    assert res["rows_added"] == len(mock_dinat_invoice_doc.items)
    assert os.path.exists(res["target_path"])

    wb = openpyxl.load_workbook(res["target_path"])
    
    # 1. Verify 3 sheets exist
    order_no = mock_dinat_invoice_doc.document_metadata.order_number
    expected_client_sheet = f"{order_no}-DATOS-CLIENTES"
    expected_prod_sheet = f"{order_no}-PRODUCTO"
    expected_deliv_sheet = f"{order_no}-ENTREGA"

    assert expected_client_sheet in wb.sheetnames
    assert expected_prod_sheet in wb.sheetnames
    assert expected_deliv_sheet in wb.sheetnames

    # 2. Check Sheet 1: DATOS-CLIENTES (11 columns)
    ws_client = wb[expected_client_sheet]
    client_headers = [cell.value for cell in ws_client[1]]
    assert client_headers == SHEET_COLUMNS["DATOS-CLIENTES"]
    assert ws_client.max_row == 2
    row2_client = [cell.value for cell in ws_client[2]]
    assert row2_client[0] == "OC-2026-08-0417"
    assert row2_client[2] == "Supermercados La Colonia, S. A. de C. V."
    assert row2_client[3] == "08019008123459"
    assert row2_client[5] == "LC-T5-TGU"
    assert row2_client[8] == 14.06793
    assert row2_client[9] == -87.194347

    # 3. Check Sheet 2: PRODUCTO (9 columns)
    ws_prod = wb[expected_prod_sheet]
    prod_headers = [cell.value for cell in ws_prod[1]]
    assert prod_headers == SHEET_COLUMNS["PRODUCTO"]
    assert ws_prod.max_row == 1 + len(mock_dinat_invoice_doc.items)
    row2_prod = [cell.value for cell in ws_prod[2]]
    assert row2_prod[0] == "OC-2026-08-0417"
    assert row2_prod[1] == "NAT-LT-MZ"
    assert row2_prod[3] == "Manzana"
    assert row2_prod[5] == 40
    assert row2_prod[6] == 960

    # 4. Check Sheet 3: ENTREGA (11 columns)
    ws_deliv = wb[expected_deliv_sheet]
    deliv_headers = [cell.value for cell in ws_deliv[1]]
    assert deliv_headers == SHEET_COLUMNS["ENTREGA"]
    assert ws_deliv.max_row == 2
    row2_deliv = [cell.value for cell in ws_deliv[2]]
    assert row2_deliv[0] == "OC-2026-08-0417"
    assert row2_deliv[1] == "José Fernando Andino Cruz"
    assert row2_deliv[2] == "0801-1992-04517"
    assert row2_deliv[4] == "DNT-1428"
    assert row2_deliv[7] == "SÍ - REALIZADA SIN PROBLEMA"


def test_append_existing_excel_sequential_lines(tmp_path, mock_dinat_invoice_doc):
    excel_file = tmp_path / "test_matriz_append_sequential.xlsx"
    file_path = str(excel_file)
    items_count = len(mock_dinat_invoice_doc.items)

    # 1. Create file first with Document 1 (Order OC-2026-08-0417)
    res1 = generate_or_append_excel([mock_dinat_invoice_doc], file_path, mode="create")
    target_path = res1["target_path"]
    
    # 2. Prepare Document 2 with a different order number (Order OC-2026-09-9999)
    doc2 = mock_dinat_invoice_doc.model_copy(deep=True)
    doc2.document_metadata.order_number = "OC-2026-09-9999"

    # 3. Append Document 2 to existing file with mode="append"
    res2 = generate_or_append_excel([doc2], target_path, mode="append")
    assert res2["success"] is True
    assert res2["rows_added"] == items_count

    # 4. Read back file and verify that Doc 1 rows are preserved intact
    wb2 = openpyxl.load_workbook(target_path)
    
    # Check that sheets contain rows for both documents
    for name in wb2.sheetnames:
        if "PRODUCTO" in name:
            ws_prod = wb2[name]
            assert ws_prod.max_row == 1 + (items_count * 2)
            assert ws_prod.cell(row=2, column=1).value == "OC-2026-08-0417"
            assert ws_prod.cell(row=2 + items_count, column=1).value == "OC-2026-09-9999"


def test_compare_columns_with_real_matriz_file():
    real_matriz_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../MATRIZ-ORDEN-COMPRA.xlsx"))
    if not os.path.exists(real_matriz_path):
        pytest.skip("MATRIZ-ORDEN-COMPRA.xlsx not present at workspace root")

    wb = openpyxl.load_workbook(real_matriz_path)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        real_headers = [cell.value for cell in ws[1]]
        if "DATOS-CLIENTES" in sheet_name:
            assert real_headers == SHEET_COLUMNS["DATOS-CLIENTES"]
        elif "PRODUCTO" in sheet_name:
            assert real_headers == SHEET_COLUMNS["PRODUCTO"]
        elif "ENTREGA" in sheet_name:
            assert real_headers == SHEET_COLUMNS["ENTREGA"]


def test_resolve_safe_excel_path_permission_fallback():
    restricted_path = r"C:\Users\usuario\non_existent_folder\test.xlsx"
    safe_path = resolve_safe_excel_path(restricted_path)

    assert safe_path != ""
    assert "exports" in safe_path or os.path.exists(os.path.dirname(safe_path))


def test_generate_or_append_excel_with_restricted_path(mock_dinat_invoice_doc):
    restricted_path = r"C:\Users\usuario\MATRIZ-ORDEN-COMPRA.xlsx"
    res = generate_or_append_excel([mock_dinat_invoice_doc], restricted_path, mode="create")

    assert res["success"] is True
    assert res["rows_added"] == len(mock_dinat_invoice_doc.items)
    assert os.path.exists(res["target_path"])
