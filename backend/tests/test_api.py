import sys
import os
import io
import pytest
import openpyxl

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from fastapi.testclient import TestClient
from main import app
from exporter import SHEET_COLUMNS


client = TestClient(app)


def test_health_endpoint():
    response = client.get("/health")
    assert response.status_code == 200
    json_data = response.json()
    assert json_data["status"] == "ok"
    assert "service" in json_data


def test_extract_pdf_mock_fallback():
    dummy_pdf = io.BytesIO(b"%PDF-1.4 mock content for extraction testing")
    files = {"file": ("test_order.pdf", dummy_pdf, "application/pdf")}

    response = client.post("/api/extract", files=files)
    assert response.status_code == 200
    json_data = response.json()
    
    assert "document_metadata" in json_data
    assert "vendor" in json_data
    assert "client" in json_data
    assert "items" in json_data
    assert "financial_totals" in json_data
    assert "transport_logistics" in json_data
    assert "delivery_status" in json_data
    assert "authorizations" in json_data
    
    # Check Sheet 1 (DATOS-CLIENTES) fields in JSON
    assert json_data["document_metadata"]["order_number"] == "OC-2026-08-0417"
    assert json_data["document_metadata"]["issue_date"] == "2026-08-17"
    assert "Supermercados La Colonia" in json_data["client"]["company_name"]
    assert json_data["client"]["rtn"] == "08019008123459"
    assert json_data["client"]["store_name"] == "T5 - La Kennedy"
    assert json_data["client"]["store_code"] == "LC-T5-TGU"
    assert json_data["client"]["coordinates"]["latitude"] == 14.06793
    assert json_data["client"]["coordinates"]["longitude"] == -87.194347

    # Check Sheet 2 (PRODUCTO) fields in JSON
    assert len(json_data["items"]) >= 1
    assert json_data["items"][0]["code"] == "NAT-LT-MZ"
    assert json_data["items"][0]["flavor"] == "Manzana"

    # Check Sheet 3 (ENTREGA) fields in JSON
    assert json_data["transport_logistics"]["driver_name"] == "José Fernando Andino Cruz"
    assert json_data["transport_logistics"]["employee_id"] == "DNT-1428"
    assert json_data["delivery_status"]["status"] == "SÍ - REALIZADA SIN PROBLEMA"


def test_export_excel_api_endpoint_three_sheets(tmp_path, mock_dinat_invoice_dict):
    target_excel = str(tmp_path / "api_export_test.xlsx")
    payload = {
        "documents": [mock_dinat_invoice_dict],
        "excel_path": target_excel,
        "mode": "create"
    }

    response = client.post("/api/export/excel", json=payload)
    assert response.status_code == 200
    json_data = response.json()
    assert json_data["success"] is True
    assert json_data["rows_added"] == len(mock_dinat_invoice_dict["items"])

    # Load workbook and check all 3 sheets
    wb = openpyxl.load_workbook(json_data["target_path"])
    sheet_names = wb.sheetnames
    assert any("DATOS-CLIENTES" in name for name in sheet_names)
    assert any("PRODUCTO" in name for name in sheet_names)
    assert any("ENTREGA" in name for name in sheet_names)

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        if "DATOS-CLIENTES" in sheet_name:
            assert headers == SHEET_COLUMNS["DATOS-CLIENTES"]
        elif "PRODUCTO" in sheet_name:
            assert headers == SHEET_COLUMNS["PRODUCTO"]
        elif "ENTREGA" in sheet_name:
            assert headers == SHEET_COLUMNS["ENTREGA"]

