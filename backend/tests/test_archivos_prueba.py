import os
import sys
import pytest
import openpyxl

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from extractor import extract_invoice_data_from_bytes, parse_pdf_text
from exporter import generate_or_append_excel, SHEET_COLUMNS
from schemas import DinatInvoiceDocument


def get_pdf_bytes(relative_path: str) -> bytes:
    full_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "../..", relative_path))
    if not os.path.exists(full_path):
        pytest.skip(f"Test file not found: {relative_path}")
    with open(full_path, "rb") as f:
        return f.read()


def test_file_01_la_granja():
    """
    Test extraction of 01_OC-2026-08-0392_LaGranja_LaColonia-T5-Kennedy_2026-08-16.pdf
    Validates:
    - Fecha de emisión: 2026-08-16
    - Hora de llegada: 06:58 a.m.
    - Hora de finalización: 07:35 a.m.
    - All products: 2 items (LGR-TP2-NJ and LGR-1L-NJ)
    - Totals, Client, Vendor, Logistics
    """
    pdf_bytes = get_pdf_bytes("archivos_prueba/01_OC-2026-08-0392_LaGranja_LaColonia-T5-Kennedy_2026-08-16.pdf")
    doc: DinatInvoiceDocument = extract_invoice_data_from_bytes(pdf_bytes, "01_OC-2026-08-0392_LaGranja_LaColonia-T5-Kennedy_2026-08-16.pdf")

    # Document Metadata
    assert doc.document_metadata.order_number == "OC-2026-08-0392"
    assert doc.document_metadata.issue_date == "2026-08-16"
    assert doc.document_metadata.dispatch_date == "2026-08-16"

    # Vendor & Client
    assert doc.vendor.brand == "LA GRANJA"
    assert "Supermercados La Colonia" in doc.client.company_name
    assert doc.client.rtn == "08019008123459"
    assert doc.client.store_name == "T5 - La Kennedy"
    assert doc.client.store_code == "LC-T5-TGU"
    assert doc.client.coordinates.latitude == 14.06793
    assert doc.client.coordinates.longitude == -87.194347

    # Line Items (All Products)
    assert len(doc.items) == 2, f"Expected 2 items, got {len(doc.items)}"
    
    item1 = doc.items[0]
    assert item1.code == "LGR-TP2-NJ"
    assert item1.flavor == "Naranja"
    assert "27" in item1.package_type
    assert item1.boxes_quantity == 85
    assert item1.total_units == 2295
    assert item1.unit_price == 195.75
    assert item1.total_amount == 16638.75

    item2 = doc.items[1]
    assert item2.code == "LGR-1L-NJ"
    assert item2.flavor == "Naranja"
    assert "12" in item2.package_type
    assert item2.boxes_quantity == 40
    assert item2.total_units == 480
    assert item2.unit_price == 372.00
    assert item2.total_amount == 14880.00

    # Financial Totals
    assert doc.financial_totals.total_boxes == 125
    assert doc.financial_totals.total_units == 2775
    assert doc.financial_totals.taxable_subtotal == 31518.75
    assert doc.financial_totals.tax_isv_15 == 4727.81
    assert doc.financial_totals.grand_total == 36246.56

    # Transport & Delivery Status
    assert "José Fernando Andino Cruz" in doc.transport_logistics.driver_name
    assert doc.transport_logistics.employee_id == "DNT-1428"
    assert doc.transport_logistics.national_id == "0801-1992-04517"
    assert doc.delivery_status.status == "SÍ - REALIZADA SIN PROBLEMA"
    assert doc.delivery_status.arrival_time == "06:58 a.m."
    assert doc.delivery_status.completion_time == "07:35 a.m."
    assert "Entrega realizada sin novedad" in doc.delivery_status.observations


def test_file_02_raptor():
    """
    Test extraction of 02_OC-2026-08-0378_Raptor_Paiz-Kennedy_2026-08-15.pdf
    Validates:
    - Fecha de emisión: 2026-08-15
    - Hora de llegada: 07:12 a.m.
    - Hora de finalización: 08:04 a.m.
    - All products: 3 items (RPT-LT355-OR, RPT-BT300-OR, RPT-BT600-OR)
    - Totals, Client, Vendor, Logistics
    """
    pdf_bytes = get_pdf_bytes("archivos_prueba/02_OC-2026-08-0378_Raptor_Paiz-Kennedy_2026-08-15.pdf")
    doc: DinatInvoiceDocument = extract_invoice_data_from_bytes(pdf_bytes, "02_OC-2026-08-0378_Raptor_Paiz-Kennedy_2026-08-15.pdf")

    # Document Metadata
    assert doc.document_metadata.order_number == "OC-2026-08-0378"
    assert doc.document_metadata.issue_date == "2026-08-15"
    assert doc.document_metadata.dispatch_date == "2026-08-15"

    # Vendor & Client
    assert doc.vendor.brand == "RAPTOR"
    assert "Operadora del Sur" in doc.client.company_name
    assert "Supermercados Paiz" in doc.client.company_name
    assert doc.client.rtn == "08019012456783"
    assert doc.client.store_name == "PZ-14 - La Kennedy"
    assert doc.client.store_code == "PZ-14-TGU"
    assert doc.client.coordinates.latitude == 14.059433
    assert doc.client.coordinates.longitude == -87.182398

    # Line Items (All Products)
    assert len(doc.items) == 3, f"Expected 3 items, got {len(doc.items)}"

    item1 = doc.items[0]
    assert item1.code == "RPT-LT355-OR"
    assert item1.flavor == "Original"
    assert "24" in item1.package_type
    assert item1.boxes_quantity == 55
    assert item1.total_units == 1320
    assert item1.unit_price == 456.00
    assert item1.total_amount == 25080.00

    item2 = doc.items[1]
    assert item2.code == "RPT-BT300-OR"
    assert item2.flavor == "Original"
    assert "12" in item2.package_type
    assert item2.boxes_quantity == 70
    assert item2.total_units == 840
    assert item2.unit_price == 192.00
    assert item2.total_amount == 13440.00

    item3 = doc.items[2]
    assert item3.code == "RPT-BT600-OR"
    assert item3.flavor == "Original"
    assert "12" in item3.package_type
    assert item3.boxes_quantity == 48
    assert item3.total_units == 576
    assert item3.unit_price == 336.00
    assert item3.total_amount == 16128.00

    # Financial Totals
    assert doc.financial_totals.total_boxes == 173
    assert doc.financial_totals.total_units == 2736
    assert doc.financial_totals.taxable_subtotal == 54648.00
    assert doc.financial_totals.tax_isv_15 == 8197.20
    assert doc.financial_totals.grand_total == 62845.20

    # Transport & Delivery Status
    assert "Elder Josué Mejía Portillo" in doc.transport_logistics.driver_name
    assert doc.transport_logistics.employee_id == "DNT-1073"
    assert doc.transport_logistics.national_id == "0801-1988-11204"
    assert doc.transport_logistics.assigned_route == "R-03 Tegucigalpa Este"
    assert "PCJ-3086" in doc.transport_logistics.transport_unit
    assert doc.delivery_status.status == "SÍ - REALIZADA SIN PROBLEMA"
    assert doc.delivery_status.arrival_time == "07:12 a.m."
    assert doc.delivery_status.completion_time == "08:04 a.m."


def test_file_03_mountain_dew():
    """
    Test extraction of 03_OC-2026-08-0361_MountainDew_MaxiDespensa-Kennedy_2026-08-14.pdf
    Validates:
    - Fecha de emisión: 2026-08-14
    - Hora de llegada: 06:05 a.m.
    - Hora de finalización: 07:48 a.m.
    - Estado de entrega: NO - ENTREGA CON INCIDENCIA
    - All products: 2 items (MTD-LT355-OR, MTD-BT600-OR)
    - Totals, Client, Vendor, Logistics
    """
    pdf_bytes = get_pdf_bytes("archivos_prueba/03_OC-2026-08-0361_MountainDew_MaxiDespensa-Kennedy_2026-08-14.pdf")
    doc: DinatInvoiceDocument = extract_invoice_data_from_bytes(pdf_bytes, "03_OC-2026-08-0361_MountainDew_MaxiDespensa-Kennedy_2026-08-14.pdf")

    # Document Metadata
    assert doc.document_metadata.order_number == "OC-2026-08-0361"
    assert doc.document_metadata.issue_date == "2026-08-14"
    assert doc.document_metadata.dispatch_date == "2026-08-14"

    # Vendor & Client
    assert doc.vendor.brand == "MOUNTAIN DEW"
    assert "Operadora del Sur" in doc.client.company_name
    assert "Maxi Despensa" in doc.client.company_name
    assert doc.client.rtn == "08019012456783"
    assert doc.client.store_name == "MD-09 - La Kennedy"
    assert doc.client.store_code == "MD-09-TGU"
    assert doc.client.coordinates.latitude == 14.061690
    assert doc.client.coordinates.longitude == -87.176625

    # Line Items (All Products)
    assert len(doc.items) == 2, f"Expected 2 items, got {len(doc.items)}"

    item1 = doc.items[0]
    assert item1.code == "MTD-LT355-OR"
    assert item1.flavor == "Original"
    assert "24" in item1.package_type
    assert item1.boxes_quantity == 62
    assert item1.total_units == 1488
    assert item1.unit_price == 300.00
    assert item1.total_amount == 18600.00

    item2 = doc.items[1]
    assert item2.code == "MTD-BT600-OR"
    assert item2.flavor == "Original"
    assert "12" in item2.package_type
    assert item2.boxes_quantity == 90
    assert item2.total_units == 1080
    assert item2.unit_price == 228.00
    assert item2.total_amount == 20520.00

    # Financial Totals
    assert doc.financial_totals.total_boxes == 152
    assert doc.financial_totals.total_units == 2568
    assert doc.financial_totals.taxable_subtotal == 39120.00
    assert doc.financial_totals.tax_isv_15 == 5868.00
    assert doc.financial_totals.grand_total == 44988.00

    # Transport & Delivery Status
    assert "Héctor Ariel Carías Flores" in doc.transport_logistics.driver_name
    assert doc.transport_logistics.employee_id == "DNT-1352"
    assert doc.transport_logistics.national_id == "0801-1990-06345"
    assert doc.transport_logistics.assigned_route == "R-01 Tegucigalpa Metropolitana"
    assert "PDL-1907" in doc.transport_logistics.transport_unit
    assert doc.delivery_status.status == "NO - ENTREGA CON INCIDENCIA"
    assert doc.delivery_status.arrival_time == "06:05 a.m."
    assert doc.delivery_status.completion_time == "07:48 a.m."
    assert "Entrega con incidencia" in doc.delivery_status.observations


@pytest.mark.parametrize("file_name,expected_order,expected_brand,expected_items_count", [
    ("archivos_prueba/04_OC-2026-08-0344_Naturas_Paiz-Miraflores_2026-08-13.pdf", "OC-2026-08-0344", "NATURAS", 12),
    ("archivos_prueba/05_OC-2026-08-0327_Raptor_LaColonia-T5-Kennedy_2026-08-12.pdf", "OC-2026-08-0327", "RAPTOR", 3),
    ("archivos_prueba/07_OC-2026-08-0296_MountainDew_Paiz-Miraflores_2026-08-10.pdf", "OC-2026-08-0296", "MOUNTAIN DEW", 2),
    ("archivos_prueba/08_OC-2026-08-0281_Naturas_LaColonia-T10-Hacienda_2026-08-09.pdf", "OC-2026-08-0281", "NATURAS", 12),
    ("archivos_prueba/09_OC-2026-08-0265_LaGranja_Paiz-Kennedy_2026-08-08.pdf", "OC-2026-08-0265", "LA GRANJA", 2),
    ("archivos_prueba/10_OC-2026-08-0248_Raptor_MaxiDespensa-Kennedy_2026-08-07.pdf", "OC-2026-08-0248", "RAPTOR", 3),
])
def test_all_remaining_archivos_prueba(file_name, expected_order, expected_brand, expected_items_count):
    pdf_bytes = get_pdf_bytes(file_name)
    doc: DinatInvoiceDocument = extract_invoice_data_from_bytes(pdf_bytes, os.path.basename(file_name))
    assert doc.document_metadata.order_number == expected_order
    assert doc.vendor.brand == expected_brand
    assert len(doc.items) == expected_items_count
    assert doc.delivery_status.arrival_time != ""
    assert doc.delivery_status.completion_time != ""


def test_excel_export_all_documents(tmp_path):
    """
    Exports documents from file 01, 02, 03 and base file together to an Excel workbook,
    verifying all 3 sheets and row counts.
    """
    files = [
        "Orden_Compra_DINAT_Naturas_LaColonia_T5_2026-08-17.pdf",
        "archivos_prueba/01_OC-2026-08-0392_LaGranja_LaColonia-T5-Kennedy_2026-08-16.pdf",
        "archivos_prueba/02_OC-2026-08-0378_Raptor_Paiz-Kennedy_2026-08-15.pdf",
        "archivos_prueba/03_OC-2026-08-0361_MountainDew_MaxiDespensa-Kennedy_2026-08-14.pdf",
    ]

    docs = []
    total_expected_items = 0
    for f in files:
        b = get_pdf_bytes(f)
        d = extract_invoice_data_from_bytes(b, os.path.basename(f))
        docs.append(d)
        total_expected_items += len(d.items)

    target_excel = str(tmp_path / "MATRIZ_ALL_ORDERS.xlsx")
    res = generate_or_append_excel(docs, target_excel, mode="create")
    assert res["success"] is True
    assert res["rows_added"] == total_expected_items

    wb = openpyxl.load_workbook(res["target_path"])
    sheet_names = wb.sheetnames
    
    client_sheet = next(s for s in sheet_names if "DATOS-CLIENTES" in s)
    prod_sheet = next(s for s in sheet_names if "PRODUCTO" in s)
    deliv_sheet = next(s for s in sheet_names if "ENTREGA" in s)

    ws_client = wb[client_sheet]
    ws_prod = wb[prod_sheet]
    ws_deliv = wb[deliv_sheet]

    # 4 client rows + 1 header
    assert ws_client.max_row == 5
    # total product rows + 1 header
    assert ws_prod.max_row == 1 + total_expected_items
    # 4 delivery rows + 1 header
    assert ws_deliv.max_row == 5
